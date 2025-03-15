from typing import Type, Generic, TypeVar, List, Generator, Optional, Union, Tuple

import pandas as pd
from openpyxl import load_workbook

from nicegui_start_project.utils.base import validate_param
from .interfaces.ExcelRow import ExcelRow

T = TypeVar("T", bound=ExcelRow)


class ExcelReader(Generic[T]):

    def __init__(self,
                 filepath: str,
                 header_var_names: Optional[List[str]] = None,
                 row_class: Optional[Type[T]] = None):
        self._filepath = filepath

        self._header_var_names = header_var_names
        self._row_class: Optional[Type[T]] = row_class

        self.__excel: Optional[pd.DataFrame] = None

        # todo: 找个关闭 self._excel 的地方

    @property
    def _excel(self):
        if self.__excel is None:
            self.__excel = load_workbook(self._filepath, data_only=True)
        return self.__excel

    def _generate_row_data(self, row: Union[List, Tuple]):
        assert self._header_var_names is not None

        res = {}
        for index, field in enumerate(self._header_var_names):
            res[field] = row[index] if index < len(row) else None  # row 如果短则补 None
        return res

    @validate_param("min_row", lambda x: x >= 1)
    def get_rows(self, min_row: int = 2, max_row: int = None) -> List[Union[Tuple, T]]:
        res = []
        for row in self._excel.active.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
            if self._row_class is None:
                elem = row
            else:
                elem = self._row_class(**self._generate_row_data(row))
            res.append(elem)
        return res
