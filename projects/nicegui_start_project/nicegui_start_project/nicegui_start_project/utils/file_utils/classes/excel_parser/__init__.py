__all__ = ["ExcelReader"]

from dataclasses import dataclass, asdict, astuple
from typing import Type, Generic, TypeVar, List, Generator, Optional, Union, Tuple

import pandas as pd
from openpyxl import load_workbook


@dataclass
class Row:
    pass


T = TypeVar("T", bound=Row)


class ExcelReader(Generic[T]):
    def __init__(self, row_class: Type[T], headers: List[str], filepath: str, min_row: int = 2):
        self.row_class = row_class
        self.headers = headers
        self.filepath = filepath
        self.min_row = min_row

        self._excel: Optional[pd.DataFrame] = None

    @property
    def excel(self):
        if self._excel is None:
            self._excel = load_workbook(self.filepath, data_only=True)
        return self._excel

    class DefaultList(list):
        def __getitem__(self, index):
            if index >= len(self):
                return None
            return super().__getitem__(index)

    @staticmethod
    def get_list_element(seq: List, index: int):
        return seq[index] if index < len(seq) else None

    def _generate_row_data(self, row: Union[List, Tuple]):
        res = {}
        # todo: row 可能短可能长... 如果短则补 None
        # assert len(row) == len(self.headers)
        for index, field in enumerate(self.headers):
            res[field] = self.get_list_element(row, index)
        return res

    def get_rows(self) -> List[T]:
        res = []
        for row in self.excel.active.iter_rows(min_row=self.min_row, values_only=True):
            res.append(self.row_class(**self._generate_row_data(row)))
        return res
